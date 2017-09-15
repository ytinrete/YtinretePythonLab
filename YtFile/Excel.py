import csv
from openpyxl import load_workbook
from openpyxl import Workbook
import xlrd


# pip3 install openpyxl
# pip3 install xlrd

def read_csv():
    try:
        res = []
        table_name = 'unknown'
        with open("../TestFile/Excel/query_result.csv") as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                obj = {}
                obj['data'] = {}
                for (k, v) in row.items():
                    pos = k.index('.')
                    table_name = k[0:pos]
                    col = k[pos + 1:]
                    obj['table'] = table_name
                    if col == 'eventtime':
                        obj['eventtime'] = v
                    elif col == 'sessionid':
                        obj['sessionid'] = v
                    elif col == 'eventname':
                        obj['eventname'] = v
                    elif col == 'test_data_deviceid':
                        obj['test_data_deviceid'] = v
                    elif col == 'appkey':
                        obj['appkey'] = v
                    elif col == 'dt':
                        obj['dt'] = v
                    else:
                        obj['data'][col] = v

                res.append(obj)

        return res, table_name
    except BaseException as e:
        print(e)


def read_xlsx():
    wb = load_workbook(filename='../TestFile/Excel/Test.xlsx')
    sheet_ranges = wb['工作表1']
    print(sheet_ranges['A2'].value)
    pass


def read_xlsx_loop(f, sheet_name, f_name, c_name):
    res = []
    wb = load_workbook(filename=f)
    for s in wb:
        if sheet_name is None or s.title.startswith(sheet_name):
            for row in s.rows:
                for cell in row:
                    # print(cell.value)
                    if str(cell.column) == c_name:
                        if type(cell.value) == str and str(cell.value).strip().startswith("C"):
                            new_block = {
                                "f_name": f_name,
                                "s_name": s.title,
                                "pos": str("[" + str(cell.column) + str(cell.row) + "]"),
                                "target": str(cell.value).strip()
                            }
                            res.append(new_block)
                        else:
                            if type(cell.value) == str:
                                print("strange rec:" + str(cell.value) + " s:" + s.title + str(
                                    "[" + str(cell.column) + str(cell.row) + "]"))
                            else:
                                if type(cell.value) == int:
                                    print("strange rec:" + str(cell.value) + " s:" + s.title + str(
                                        "[" + str(cell.column) + str(cell.row) + "]"))

    return res


def write_to_xlsx(f1_name, f2_name, same):
    wb = Workbook()
    ws = wb.active
    for i in range(0, len(same) - 1):
        item = same[i]
        index = i + 1
        if index == 1:
            ws.cell(row=index, column=1).value = "target"
            ws.cell(row=index, column=2).value = "pos A"
            ws.cell(row=index, column=3).value = "pos B"
            ws.cell(row=index, column=4).value = "count:" + str(len(same))
        else:
            ws.cell(row=index, column=1).value = item["A"]["target"]
            ws.cell(row=index, column=2).value = \
                "file[" + item["A"]["f_name"] + "]sheet[" + item["A"]["s_name"] + "]pos" + item["A"]["pos"]
            ws.cell(row=index, column=3).value = \
                "file[" + item["B"]["f_name"] + "]sheet[" + item["B"]["s_name"] + "]pos" + item["B"]["pos"]
    wb.save(str("compare_excel/" + f1_name + "and" + f2_name + "equal.xlsx"))


def read_xls():
    wb = xlrd.open_workbook('../TestFile/Excel/Test.xls')
    sh = wb.sheet_by_index(0)
    # print(sh.cell(2, 2).value)
    print(sh.row_values(1)[0])
    pass


if __name__ == '__main__':
    cvs_list, name = read_csv()

    for item in cvs_list:
        print(item)

    read_xlsx()

    read_xls()

    pass
